import io
import logging

from odoo import _, api, fields, models

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger = logging.getLogger(__name__)
    _logger.warning("openpyxl is not installed. XLSX export will not work.")

_logger = logging.getLogger(__name__)


class OdfeReportXlsx(models.AbstractModel):
    _name = 'odfe.report.xlsx'
    _description = 'XLSX Report Generator'

    HEADER_FILL = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    SUBHEADER_FONT = Font(name='Calibri', bold=True, size=11)
    DATA_FONT = Font(name='Calibri', size=10)
    CURRENCY_FORMAT = '#,##0.00'
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    def generate_sales_xlsx(self, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sales Report'
        self._write_title(ws, 'Sales Report', data)
        ws.append([])

        summary = data.get('summary', {})
        if summary:
            ws.append(['Summary'])
            ws.cell(row=ws.max_row, column=1).font = self.SUBHEADER_FONT
            ws.cell(row=ws.max_row, column=1).fill = self.SUBHEADER_FILL

            summary_fields = [
                ('Total Revenue', 'total_revenue'),
                ('Total Orders', 'total_orders'),
                ('Total Discounts', 'total_discounts'),
                ('Total Taxes', 'total_taxes'),
                ('Avg Order Value', 'average_order_value'),
                ('Cancelled Orders', 'cancelled_orders'),
            ]
            for label, key in summary_fields:
                if key in summary:
                    row_num = ws.max_row + 1
                    ws.cell(row=row_num, column=1, value=label).font = self.DATA_FONT
                    cell = ws.cell(row=row_num, column=2, value=summary[key])
                    cell.font = self.DATA_FONT
                    if key != 'total_orders' and key != 'cancelled_orders':
                        cell.number_format = self.CURRENCY_FORMAT
                    self._apply_border(ws, row_num, 1, 2)

        breakdown = data.get('breakdown', [])
        if breakdown:
            ws.append([])
            ws.append(['Daily Breakdown'])
            ws.cell(row=ws.max_row, column=1).font = self.SUBHEADER_FONT
            ws.cell(row=ws.max_row, column=1).fill = self.SUBHEADER_FILL
            self._write_table(ws, breakdown, ['Label', 'Revenue', 'Orders', 'Discounts', 'Taxes'],
                              ['label', 'revenue', 'orders', 'discounts', 'taxes'],
                              currency_fields=['revenue', 'discounts', 'taxes'])

        self._auto_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()
        buf.close()
        return xlsx

    def generate_report_xlsx(self, report_type, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        title_map = {
            'sales': 'Sales Report',
            'orders': 'Orders Report',
            'revenue': 'Revenue Report',
            'products': 'Products Report',
            'employees': 'Employee Performance Report',
            'sessions': 'Session Report',
        }
        title = title_map.get(report_type, 'Report')
        ws.title = title[:31]
        self._write_title(ws, title, data)

        summary = data.get('summary', {})
        if summary:
            ws.append([])
            ws.append(['Summary'])
            ws.cell(row=ws.max_row, column=1).font = self.SUBHEADER_FONT
            ws.cell(row=ws.max_row, column=1).fill = self.SUBHEADER_FILL
            for key, val in summary.items():
                if isinstance(val, list):
                    continue
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=1, value=key.replace('_', ' ').title()).font = self.DATA_FONT
                cell = ws.cell(row=row_num, column=2, value=val)
                cell.font = self.DATA_FONT
                if isinstance(val, float):
                    cell.number_format = self.CURRENCY_FORMAT
                    self._apply_border(ws, row_num, 1, 2)

        breakdown = data.get('breakdown', data.get('data', []))
        if breakdown and isinstance(breakdown, list) and breakdown:
            ws.append([])
            keys = list(breakdown[0].keys())
            labels = [k.replace('_', ' ').title() for k in keys]
            currency_fields = [keys[i] for i, k in enumerate(keys)
                               if breakdown[0].get(k) is not None and isinstance(breakdown[0][k], float)]
            self._write_table(ws, breakdown, labels, keys, currency_fields=currency_fields)

        self._auto_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()
        buf.close()
        return xlsx

    def _write_title(self, ws, title, data):
        ws.cell(row=1, column=1, value=title).font = Font(name='Calibri', bold=True, size=14, color='2E5090')
        ws.cell(row=2, column=1, value=f"Company: {self.env.company.name}").font = self.DATA_FONT
        ws.cell(row=2, column=2, value=f"Generated: {fields.Datetime.now()}").font = self.DATA_FONT
        if data.get('date_from'):
            ws.cell(row=3, column=1, value=f"Period: {data.get('date_from', '')} - {data.get('date_to', '')}").font = self.DATA_FONT

    def _write_table(self, ws, data, labels, keys, currency_fields=None):
        currency_fields = currency_fields or []
        header_row = ws.max_row + 1
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            self.THIN_BORDER

        for row_idx, record in enumerate(data, header_row + 1):
            for col_idx, key in enumerate(keys, 1):
                val = record.get(key, '')
                if val is None:
                    val = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER
                if key in currency_fields:
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.alignment = Alignment(horizontal='right')
                elif isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

            if row_idx % 2 == 0:
                fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                for col_idx in range(1, len(keys) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    def _apply_border(self, ws, row, col_from, col_to):
        for col in range(col_from, col_to + 1):
            ws.cell(row=row, column=col).border = self.THIN_BORDER

    def _auto_width(self, ws):
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
            adjusted = min(max_length + 3, 40)
            ws.column_dimensions[col_letter].width = adjusted
